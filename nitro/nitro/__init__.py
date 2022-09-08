__version__ = "0.3.1"

import pysam
import pandas
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

position_column = [
    "",
    "",
    "Position",
    241,
    670,
    2790,
    2832,
    3037,
    4184,
    4321,
    5389,
    8393,
    9344,
    9424,
    9534,
    9866,
    10029,
    10198,
    10447,
    10449,
    11537,
    12880,
    13195,
    14408,
    15240,
    15714,
    17410,
    18163,
    19955,
    20055,
    21618,
    21762,
    21846,
    21987,
    22200,
    22578,
    22673,
    22674,
    22679,
    22686,
    22688,
    22775,
    22786,
    22813,
    22882,
    22992,
    22995,
    23013,
    23040,
    23048,
    23055,
    23063,
    23075,
    23202,
    23403,
    23525,
    23599,
    23604,
    23854,
    23948,
    24130,
    24424,
    24469,
    24503,
    25000,
    25584,
    26060,
    26270,
    26530,
    26577,
    26709,
    26858,
    26259,
    27382,
    27383,
    27384,
    27807,
    28271,
    28311,
    28881,
    28882,
    28883,
    29510,
]
template_df = pandas.DataFrame(
    data=[
        [
            "",
            "",
            "Reference",
            "C",
            "T",
            "C",
            "A",
            "C",
            "G",
            "C",
            "T",
            "G",
            "C",
            "A",
            "C",
            "C",
            "C",
            "C",
            "G",
            "C",
            "A",
            "C",
            "T",
            "C",
            "C",
            "C",
            "C",
            "A",
            "C",
            "A",
            "C",
            "C",
            "C",
            "G",
            "T",
            "G",
            "T",
            "C",
            "T",
            "C",
            "A",
            "G",
            "A",
            "G",
            "T",
            "G",
            "C",
            "A",
            "A",
            "G",
            "A",
            "A",
            "T",
            "C",
            "A",
            "C",
            "T",
            "C",
            "C",
            "G",
            "C",
            "A",
            "T",
            "C",
            "C",
            "C",
            "C",
            "C",
            "A",
            "C",
            "G",
            "C",
            "A",
            "G",
            "A",
            "T",
            "C",
            "A",
            "C",
            "G",
            "G",
            "G",
            "A",
        ],
        [
            "",
            "",
            "Omicron / BA.1 / 21K",
            "T",
            "T",
            "C",
            "G",
            "T",
            "G",
            "C",
            "G",
            "A",
            "C",
            "A",
            "C",
            "C",
            "T",
            "C",
            "G",
            "A",
            "G",
            "C",
            "C",
            "T",
            "T",
            "C",
            "C",
            "G",
            "C",
            "A",
            "C",
            "T",
            "T",
            "G",
            "T",
            "A",
            "C",
            "T",
            "C",
            "T",
            "A",
            "G",
            "A",
            "G",
            "T",
            "A",
            "A",
            "C",
            "G",
            "A",
            "G",
            "T",
            "C",
            "A",
            "G",
            "T",
            "G",
            "A",
            "A",
            "T",
            "A",
            "T",
            "A",
            "T",
            "T",
            "T",
            "C",
            "T",
            "G",
            "G",
            "A",
            "C",
            "C",
            "G",
            "A",
            "T",
            "T",
            "T",
            "T",
            "A",
            "A",
            "C",
            "A",
        ],
        [
            "",
            "",
            "Omicron / BA.2 / 21L",
            "T",
            "G",
            "T",
            "A",
            "T",
            "A",
            "T",
            "T",
            "G",
            "T",
            "G",
            "T",
            "T",
            "T",
            "T",
            "A",
            "A",
            "A",
            "T",
            "T",
            "T",
            "C",
            "T",
            "T",
            "G",
            "T",
            "G",
            "T",
            "C",
            "C",
            "A",
            "G",
            "A",
            "T",
            "T",
            "C",
            "T",
            "G",
            "A",
            "C",
            "T",
            "G",
            "A",
            "A",
            "C",
            "G",
            "G",
            "G",
            "T",
            "C",
            "C",
            "G",
            "T",
            "G",
            "A",
            "A",
            "T",
            "C",
            "T",
            "A",
            "C",
            "T",
            "T",
            "T",
            "T",
            "A",
            "G",
            "A",
            "T",
            "C",
            "C",
            "T",
            "C",
            "T",
            "T",
            "T",
            "A",
            "A",
            "C",
            "C",
        ],
    ],
    columns=position_column,
)
row_df = pandas.DataFrame(columns=position_column)


bamfile = pysam.AlignmentFile(
    "/home/nsm/Desktop/All_Development/Nitro/nitro/nitro/workflow/output/Sample2/Sample2_merged.bam",
    "rb",
)
reference = bamfile.references[0]


##############################################################################################

wb = Workbook()
ws = wb.active
ws.title = "BA.1 X BA.2"

for r in dataframe_to_rows(template_df, index=False, header=True):
    ws.append(r)


for rows in ws.iter_rows(min_row=1, max_row=4, min_col=4):
    for cell in rows:
        if cell.row == 1:
            cell.alignment = Alignment(
                textRotation=90, vertical="center", horizontal="center"
            )
        else:
            cell.alignment = Alignment(vertical="center", horizontal="center")

        if cell.row == 3:
            cell.fill = PatternFill(
                start_color="ef4444", end_color="ef4444", fill_type="solid"
            )
        elif cell.row == 4:
            prev_cell = ws.cell(3, cell.col_idx)
            first_value = prev_cell.value
            second_value = cell.value
            if first_value == second_value:
                prev_cell.fill = PatternFill(
                    start_color="78716c", end_color="78716c", fill_type="solid"
                )
                cell.fill = PatternFill(
                    start_color="78716c", end_color="78716c", fill_type="solid"
                )
            else:
                cell.fill = PatternFill(
                    start_color="22c55e", end_color="22c55e", fill_type="solid"
                )

row_df.loc[0] = [
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
]

row_df.loc[1] = [
    "Assessment",
    "Value",
    "MH-INSACOG-CSIR-NEERI947",
    "T",
    "T",
    "C",
    "G",
    "T",
    "G",
    "C",
    "G",
    "A",
    "C",
    "A",
    "C",
    "C",
    "T",
    "C",
    "G",
    "A",
    "G",
    "C",
    "C",
    "T",
    "T",
    "C",
    "C",
    "G",
    "T",
    "G",
    "T",
    "C",
    "C",
    "A",
    "G",
    "A",
    "T",
    "T",
    "C",
    "T",
    "G",
    "A",
    "C",
    "T",
    "G",
    "A",
    "A",
    "C",
    "G",
    "G",
    "G",
    "T",
    "C",
    "C",
    "G",
    "T",
    "G",
    "A",
    "A",
    "T",
    "C",
    "T",
    "A",
    "C",
    "T",
    "T",
    "T",
    "T",
    "A",
    "G",
    "A",
    "T",
    "C",
    "C",
    "T",
    "C",
    "T",
    "T",
    "T",
    "A",
    "A",
    "C",
    "C",
]

row_df.loc[2] = [
    "Avg read at defining sites",
    7.225,
    "Total reads",
    3,
    15,
    11,
    12,
    5,
    5,
    19,
    6,
    22,
    11,
    5,
    4,
    9,
    6,
    7,
    3,
    3,
    10,
    13,
    25,
    3,
    5,
    9,
    10,
    7,
    10,
    4,
    13,
    6,
    1,
    1,
    6,
    1,
    4,
    4,
    4,
    4,
    4,
    3,
    3,
    6,
    2,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    8,
    7,
    10,
    21,
    21,
    6,
    7,
    8,
    4,
    7,
    12,
    17,
    6,
    5,
    8,
    12,
    7,
    8,
    24,
    9,
    1,
    1,
    1,
    20,
    14,
    14,
    3,
    3,
    3,
    7,
]

row_df.loc[3] = [
    "Max read",
    25,
    "A",
    0,
    0,
    0,
    7,
    0,
    0,
    0,
    0,
    17,
    0,
    4,
    0,
    0,
    0,
    0,
    2,
    2,
    5,
    0,
    0,
    0,
    0,
    0,
    0,
    2,
    0,
    3,
    0,
    0,
    0,
    1,
    1,
    1,
    0,
    0,
    0,
    0,
    2,
    2,
    1,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    5,
    0,
    0,
    0,
    20,
    5,
    0,
    7,
    0,
    7,
    0,
    0,
    0,
    0,
    0,
    6,
    0,
    8,
    0,
    0,
    0,
    1,
    0,
    0,
    0,
    0,
    3,
    3,
    0,
    5,
]

row_df.loc[4] = [
    "Min read",
    0,
    "C",
    0,
    0,
    5,
    0,
    0,
    0,
    17,
    0,
    0,
    7,
    0,
    4,
    6,
    1,
    5,
    0,
    1,
    0,
    5,
    6,
    0,
    2,
    2,
    6,
    0,
    5,
    0,
    6,
    4,
    1,
    0,
    0,
    0,
    2,
    0,
    4,
    0,
    0,
    0,
    2,
    0,
    1,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    3,
    0,
    0,
    0,
    0,
    0,
    0,
    1,
    0,
    0,
    5,
    0,
    0,
    3,
    0,
    0,
    0,
    0,
    9,
    0,
    0,
    0,
    0,
    0,
    0,
    1,
    0,
    0,
    3,
    2,
]

row_df.loc[5] = [
    "No of Breakpoints",
    15,
    "G",
    0,
    13,
    0,
    5,
    0,
    5,
    0,
    2,
    5,
    0,
    1,
    0,
    0,
    0,
    0,
    1,
    0,
    5,
    0,
    0,
    0,
    0,
    0,
    0,
    5,
    0,
    1,
    1,
    0,
    0,
    0,
    2,
    0,
    0,
    0,
    0,
    0,
    2,
    1,
    0,
    0,
    1,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    7,
    0,
    21,
    0,
    1,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    6,
    7,
    0,
    0,
    0,
    1,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
]

row_df.loc[6] = [
    "No of contaminated sites",
    6,
    "T",
    3,
    2,
    6,
    0,
    5,
    0,
    2,
    4,
    0,
    4,
    0,
    0,
    3,
    5,
    2,
    0,
    0,
    0,
    8,
    19,
    3,
    3,
    7,
    4,
    0,
    5,
    0,
    6,
    2,
    0,
    0,
    3,
    0,
    2,
    4,
    0,
    4,
    0,
    0,
    0,
    6,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    0,
    10,
    0,
    1,
    0,
    7,
    0,
    4,
    0,
    7,
    17,
    6,
    2,
    8,
    0,
    0,
    0,
    15,
    9,
    0,
    0,
    1,
    20,
    14,
    13,
    0,
    0,
    0,
    0,
]

row_df.loc[7] = [
    "Sequence quality",
    "Low Depth",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
]

print(row_df)

for r in dataframe_to_rows(row_df, index=False, header=False):
    ws.append(r)

for rows in ws.iter_rows(min_row=5, min_col=4):
    for cell in rows:
        cell.alignment = Alignment(vertical="center", horizontal="center")

ws.row_dimensions[1].height = 50
ws.freeze_panes = "D5"
wb.save("Test_sheet.xlsx")
wb.close()
